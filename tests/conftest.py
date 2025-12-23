"""
pytest 配置文件和共用 fixtures
"""
import os
import pytest
import shutil
from pathlib import Path
from fastapi.testclient import TestClient
import sys

# 將專案根目錄加入 Python 路徑
sys.path.insert(0, str(Path(__file__).parent.parent))

import main
from main import app, file_lock_manager

# 測試用的 API Token
TEST_TOKEN = "test-token-12345"

@pytest.fixture(scope="session")
def test_data_dir():
    """建立測試資料目錄"""
    test_dir = Path("tests/test_data")
    test_dir.mkdir(parents=True, exist_ok=True)
    yield test_dir
    # 測試結束後清理（忽略權限錯誤）
    if test_dir.exists():
        import time
        for attempt in range(3):
            try:
                shutil.rmtree(test_dir)
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.2)
                # 最後一次嘗試失敗就放棄（不影響測試結果）

@pytest.fixture(scope="function")
def clean_test_env(test_data_dir, monkeypatch):
    """每個測試前清理環境"""
    # 直接修改 main 模組中的變數（因為環境變數在模組載入時就已讀取）
    monkeypatch.setattr(main, "API_TOKEN", TEST_TOKEN)
    monkeypatch.setattr(main, "EXCEL_ROOT_DIR", test_data_dir)
    
    # 清理測試資料目錄（帶重試邏輯）
    import time
    for file in test_data_dir.glob("*.xlsx"):
        for attempt in range(3):
            try:
                file.unlink()
                break
            except PermissionError:
                if attempt < 2:
                    time.sleep(0.1)
                # 最後一次嘗試失敗就放棄
    
    # 清理所有鎖定
    file_lock_manager.locks.clear()
    file_lock_manager.lock_times.clear()
    
    yield test_data_dir

@pytest.fixture(scope="function")
def client(clean_test_env):
    """建立測試客戶端"""
    with TestClient(app) as test_client:
        yield test_client

@pytest.fixture
def auth_headers():
    """認證標頭"""
    return {"Authorization": f"Bearer {TEST_TOKEN}"}

@pytest.fixture
def sample_excel_file(clean_test_env):
    """建立範例 Excel 檔案"""
    import openpyxl
    
    file_path = clean_test_env / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    
    # 建立表頭
    headers = ["ID", "Name", "Department", "Salary"]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # 建立測試資料
    test_data = [
        ["E001", "John Doe", "Engineering", 75000],
        ["E002", "Jane Smith", "Sales", 65000],
        ["E003", "Bob Johnson", "HR", 60000],
    ]
    
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(file_path)
    wb.close()
    
    yield file_path
    
    # 清理由 clean_test_env 負責，這裡不需要刪除
